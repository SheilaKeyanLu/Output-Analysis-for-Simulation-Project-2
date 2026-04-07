"""
paired_ttest.py
---------------
Compute paired t-test (baseline vs improved) for all 4 KPIs:
  1. Average Cycle Time
  2. Order Rejection Rate
  3. Avg Waiting Time - Aluminium Block
  4. Avg Waiting Time - Hard Disk

Baseline data : data/ALL_KPI_SUMMARY.xlsx
Improved data : data/CycleTime.xlsx  (col1=timestamp, col2=cycle time)
                data/Rejection.xlsx  (col1=rejection rate, one row per sheet)
                data/Wait.xlsx       (col1=Al wait, col2=HD wait, one row per sheet)

Pairing is done by seed (sheet name). Only seeds present in BOTH
baseline and improved are used.
"""

import os, re
import numpy as np
import pandas as pd
import openpyxl
from scipy import stats

# ── paths ──────────────────────────────────────────────────────────────────
DIR = os.path.dirname(os.path.abspath(__file__))
BASELINE_DATA = os.path.join(DIR, "baseline")
IMPROVEMENT_DATA = os.path.join(DIR, "improvement1_data")
                                
BASELINE_FILE  = os.path.join(BASELINE_DATA, "ALL_KPI_SUMMARY.xlsx")
IMP_CYCLE_FILE = os.path.join(IMPROVEMENT_DATA, "CycleTime.xlsx")
IMP_REJ_FILE   = os.path.join(IMPROVEMENT_DATA, "Rejection.xlsx")
IMP_WAIT_FILE  = os.path.join(IMPROVEMENT_DATA, "Wait.xlsx")

WARMUP   = 7400    # minutes
END_TIME = 14800   # minutes
ALPHA    = 0.05


# ══════════════════════════════════════════════════════════════════════════════
# BASELINE READERS
# ══════════════════════════════════════════════════════════════════════════════

def read_baseline_cycletime(path, warmup=WARMUP, end_time=END_TIME):
    """
    Reads 'AverageCycleTime' sheet of ALL_KPI_SUMMARY.xlsx.
    Column format: R<n>_TIME(SEED=<s>)  /  R<n>_CYCLE(SEED=<s>)
    Filters to [warmup, end_time] window, returns {seed: mean_cycle_time}.
    """
    df = pd.read_excel(path, sheet_name="AverageCycleTime", header=0)
    time_pat  = re.compile(r"R\d+_TIME\(SEED=(\d+)\)")
    cycle_pat = re.compile(r"R\d+_CYCLE\(SEED=(\d+)\)")

    col_map = {}
    for col in df.columns:
        col_str = str(col).strip()
        m = time_pat.fullmatch(col_str)
        if m:
            col_map.setdefault(int(m.group(1)), {})["TIME"] = col
            continue
        m = cycle_pat.fullmatch(col_str)
        if m:
            col_map.setdefault(int(m.group(1)), {})["CYCLE"] = col

    result = {}
    for seed, cols in col_map.items():
        if "TIME" not in cols or "CYCLE" not in cols:
            continue
        time_s  = pd.to_numeric(df[cols["TIME"]],  errors="coerce")
        cycle_s = pd.to_numeric(df[cols["CYCLE"]], errors="coerce")
        mask = time_s.notna() & cycle_s.notna() & \
               (time_s >= warmup) & (time_s <= end_time)
        selected = cycle_s[mask]
        if len(selected) > 0:
            result[seed] = float(selected.mean())
    return result   # {seed: mean}


def read_baseline_rejection(path):
    """
    Reads 'RejectionData' sheet (header row = row 2).
    Returns {seed: rejection_rate}.
    """
    df = pd.read_excel(path, sheet_name="RejectionData", header=1)
    df.columns = ["SEED", "Rejection_Rate", *df.columns[2:]]
    result = {}
    for _, row in df.iterrows():
        try:
            result[int(row["SEED"])] = float(row["Rejection_Rate"])
        except (ValueError, TypeError):
            pass
    return result


def read_baseline_wait(path):
    """
    Reads 'WaitData' sheet (header row = row 2).
    Returns {seed: (al_wait, hd_wait)}.
    """
    df = pd.read_excel(path, sheet_name="WaitData", header=1)
    df.columns = ["SEED", "AluminumBlock", "HardDisk", *df.columns[3:]]
    result = {}
    for _, row in df.iterrows():
        try:
            result[int(row["SEED"])] = (float(row["AluminumBlock"]),
                                        float(row["HardDisk"]))
        except (ValueError, TypeError):
            pass
    return result


# ══════════════════════════════════════════════════════════════════════════════
# IMPROVED READERS  (general: works for any number of sheets)
# ══════════════════════════════════════════════════════════════════════════════

def read_improved_cycletime(path, warmup=WARMUP, end_time=END_TIME):
    """
    Each sheet name = seed (integer).
    Col 0 = completion timestamp,  Col 1 = cycle time.
    Filters to [warmup, end_time] window, returns {seed: mean_cycle_time}.
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    result = {}
    for sheet in wb.sheetnames:
        try:
            seed = int(sheet)
        except ValueError:
            continue
        ws = wb[sheet]
        times, cycles = [], []
        for row in ws.iter_rows(values_only=True):
            if (len(row) >= 2
                    and isinstance(row[0], (int, float))
                    and isinstance(row[1], (int, float))):
                times.append(float(row[0]))
                cycles.append(float(row[1]))
        selected = [c for t, c in zip(times, cycles)
                    if warmup <= t <= end_time]
        if selected:
            result[seed] = float(np.mean(selected))
    wb.close()
    return result


def read_improved_rejection(path):
    """
    Each sheet name = seed. Col 0 = rejection rate (single row per sheet).
    Returns {seed: rejection_rate}.
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    result = {}
    for sheet in wb.sheetnames:
        try:
            seed = int(sheet)
        except ValueError:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            if isinstance(row[0], (int, float)):
                result[seed] = float(row[0])
                break
    wb.close()
    return result


def read_improved_wait(path):
    """
    Each sheet name = seed. Col 0 = Al wait, Col 1 = HD wait (single row).
    Returns {seed: (al_wait, hd_wait)}.
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    result = {}
    for sheet in wb.sheetnames:
        try:
            seed = int(sheet)
        except ValueError:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            if (len(row) >= 2
                    and isinstance(row[0], (int, float))
                    and isinstance(row[1], (int, float))):
                result[seed] = (float(row[0]), float(row[1]))
                break
    wb.close()
    return result


# ══════════════════════════════════════════════════════════════════════════════
# PAIRED t-TEST
# ══════════════════════════════════════════════════════════════════════════════

def paired_ttest(baseline_dict, improved_dict, alpha=ALPHA, label="KPI"):
    """
    Compute two-sided paired t-test on D_i = baseline_i - improved_i.
    Returns a result dict.
    """
    common = sorted(set(baseline_dict) & set(improved_dict))
    if len(common) < 2:
        print(f"[{label}] WARNING: only {len(common)} common seeds – skipping.")
        return None

    b   = np.array([baseline_dict[s] for s in common], dtype=float)
    imp = np.array([improved_dict[s] for s in common], dtype=float)
    d   = b - imp                          # positive = baseline worse

    n       = len(d)
    D_bar   = d.mean()
    S_D     = d.std(ddof=1)
    SE      = S_D / np.sqrt(n)
    t_crit  = stats.t.ppf(1 - alpha / 2, df=n - 1)
    t_stat  = D_bar / SE
    p_val   = 2 * stats.t.sf(abs(t_stat), df=n - 1)
    hw      = t_crit * SE

    # per-system summary
    b_mean, b_sd   = b.mean(),   b.std(ddof=1)
    imp_mean, imp_sd = imp.mean(), imp.std(ddof=1)
    b_hw   = t_crit * b_sd   / np.sqrt(n)
    imp_hw = t_crit * imp_sd / np.sqrt(n)

    return {
        "label":        label,
        "seeds":        common,
        "n":            n,
        # baseline
        "base_mean":    b_mean,
        "base_sd":      b_sd,
        "base_ci":      (b_mean - b_hw, b_mean + b_hw),
        "base_hw":      b_hw,
        # improved
        "imp_mean":     imp_mean,
        "imp_sd":       imp_sd,
        "imp_ci":       (imp_mean - imp_hw, imp_mean + imp_hw),
        "imp_hw":       imp_hw,
        # paired diff
        "D_bar":        D_bar,
        "S_D":          S_D,
        "t_stat":       t_stat,
        "diff_ci":      (D_bar - hw, D_bar + hw),
        "diff_hw":      hw,
        "p_value":      p_val,
    }


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("Reading baseline data …")
    base_cycle = read_baseline_cycletime(BASELINE_FILE)
    base_rej   = read_baseline_rejection(BASELINE_FILE)
    base_wait  = read_baseline_wait(BASELINE_FILE)

    print("Reading improved data …")
    imp_cycle  = read_improved_cycletime(IMP_CYCLE_FILE)
    imp_rej    = read_improved_rejection(IMP_REJ_FILE)
    imp_wait   = read_improved_wait(IMP_WAIT_FILE)

    # Split wait into Al and HD dicts
    base_al  = {s: v[0] for s, v in base_wait.items()}
    base_hd  = {s: v[1] for s, v in base_wait.items()}
    imp_al   = {s: v[0] for s, v in imp_wait.items()}
    imp_hd   = {s: v[1] for s, v in imp_wait.items()}

    print("\n" + "═"*70)

    results = []
    for label, b_dict, i_dict in [
        ("Average Cycle Time (min)",          base_cycle, imp_cycle),
        ("Order Rejection Rate",               base_rej,   imp_rej),
        ("Avg Wait Time – Aluminium (min)",    base_al,    imp_al),
        ("Avg Wait Time – Hard Disk (min)",    base_hd,    imp_hd),
    ]:
        r = paired_ttest(b_dict, i_dict, label=label)
        if r is None:
            continue
        results.append(r)

        print(f"\n{'─'*60}")
        print(f"KPI: {r['label']}")
        print(f"  Common seeds (n={r['n']}): {r['seeds']}")
        print(f"\n  BASELINE  : mean={r['base_mean']:.4f}  SD={r['base_sd']:.4f}"
              f"  95%CI=({r['base_ci'][0]:.4f}, {r['base_ci'][1]:.4f})"
              f"  HW={r['base_hw']:.4f}")
        print(f"  IMPROVED  : mean={r['imp_mean']:.4f}  SD={r['imp_sd']:.4f}"
              f"  95%CI=({r['imp_ci'][0]:.4f}, {r['imp_ci'][1]:.4f})"
              f"  HW={r['imp_hw']:.4f}")
        print(f"\n  Paired diff (B-I):")
        print(f"    D̄  = {r['D_bar']:.4f}")
        print(f"    S_D = {r['S_D']:.4f}")
        print(f"    t   = {r['t_stat']:.4f}")
        print(f"    95%CI for μ_D = ({r['diff_ci'][0]:.4f}, {r['diff_ci'][1]:.4f})")
        p = r['p_value']
        p_str = f"{p:.4e}" if p < 0.0001 else f"{p:.4f}"
        print(f"    p-value = {p_str}")

    # ── LaTeX table snippets ─────────────────────────────────────────────────
    print("\n\n" + "═"*70)
    print("LaTeX: Table — Baseline vs Improved")
    print("═"*70)
    for r in results:
        lbl = r["label"]
        print(f"\\multicolumn{{6}}{{l}}{{\\textit{{{lbl}}}}} \\\\")
        print(f"  & Baseline & {r['base_mean']:.2f} & {r['base_sd']:.2f}"
              f" & ({r['base_ci'][0]:.2f},\\ {r['base_ci'][1]:.2f})"
              f" & {r['base_hw']:.2f} \\\\")
        print(f"  & Improved & {r['imp_mean']:.2f} & {r['imp_sd']:.2f}"
              f" & ({r['imp_ci'][0]:.2f},\\ {r['imp_ci'][1]:.2f})"
              f" & {r['imp_hw']:.2f} \\\\[2pt]")

    print("\n\n" + "═"*70)
    print("LaTeX: Table — Paired t-Test Results")
    print("═"*70)
    for r in results:
        p = r["p_value"]
        p_str = "$<0.001$" if p < 0.001 else f"${p:.4f}$"
        print(
            f"{r['label']} & {r['D_bar']:.2f} & {r['S_D']:.2f}"
            f" & {r['t_stat']:.1f}"
            f" & ({r['diff_ci'][0]:.2f},\\ {r['diff_ci'][1]:.2f})"
            f" & {p_str} \\\\"
        )


if __name__ == "__main__":
    main()
